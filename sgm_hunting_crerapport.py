# -*- coding: utf-8 -*-

"""
    ***************************************************************************
    * Plugin name:   SgmHunting
    * Plugin type:   QGIS 3 plugin
    * Module:        CreRapport class
    * Description:   Class to do the job for creating XL lists
    *                the report
    * Specific lib:  openpyxl
    * First release: 2023-09-05
    * Last release:  2023-09-07
    * Copyright:     (C)2023 SIGMOE
    * Email:         em at sigmoe.fr
    * License:       GPL v3
    ***************************************************************************

    ***************************************************************************
    * This program is free software: you can redistribute it and/or modify
    * it under the terms of the GNU General Public License as published by
    * the Free Software Foundation, either version 3 of the License, or
    * (at your option) any later version.
    ***************************************************************************
"""


from qgis.PyQt import uic
from qgis.PyQt.QtCore import Qt, pyqtSignal
from qgis.PyQt.QtWidgets import QMessageBox, QWidget, qApp
from qgis.PyQt.QtGui import QTextCursor
from qgis.core import QgsExpression, QgsFeatureRequest

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (NamedStyle,
                             PatternFill,
                             Border,
                             Side,
                             Alignment,
                             Font)

import os.path

from .sgm_hunting_globalvars import *
from .sgm_hunting_globalfnc import *


gui_dlg_hunting_msg, _ = uic.loadUiType(
        os.path.join(os.path.dirname(__file__), r"gui/dlg_hunting_msg.ui"))


class ExportXl:
    
    def __init__(self, iface, canvas, project, plugin_dir, conf):

        self.iface = iface
        self.canvas = canvas
        self.project = project
        self.plugin_dir = plugin_dir
        self.conf = conf
        for k, v in self.conf.items():
            self.__dict__[k[:-4]] = v
        

    def launch(self) :
        # Prepare the msg window
        self.feedback = MsgWnd()
        # Capture the signal to launch the process
        self.feedback.send_ok.connect(self.ok_param) 
        # Modal window
        self.feedback.setWindowModality(Qt.ApplicationModal)
        # Show the parameters window
        self.feedback.show()


    def ok_param(self):
        '''
            Launch the process once the msg window is validated
        '''
        # Starts the export only if the new chasse layers exists
        parclot_lyr = self.project.mapLayersByName(self.parclot_lyrname)[0]
        nwlot_lyr = self.project.mapLayersByName(self.nwlots_lyrname)[0]
        if not parclot_lyr.isValid() or not nwlot_lyr.isValid():
            QMessageBox.warning(
                                    self.feedback, 
                                    alert_lyr_msg_txt[0], 
                                    alert_lyr_msg_txt[1].format(parc_lyrname, self.lot_lyrname)
                                    )
        else:
            # Finds the list of Lots
            lot_lst = get_fldval_sorted(nwlot_lyr, self.lot_attname)
            # Initializes progressbar
            self.feedback.pg_bar.setMaximum(1 + len(lot_lst))
            self.pgb_val = 0
            xl_filepath = os.path.join(self.export_rep, f"{self.export_fname}.xlsx")
            try: 
                self.send_msg(exportxl_msg_txt[0])
                wb = Workbook()
                wb.remove(wb.active)
                self.ws = None
                # One XL sheet per lot
                for lot in lot_lst:
                    exp_s = "\"" + self.lot_attname + "\"= '" + lot + "'"
                    p_exp = QgsExpression(exp_s)
                    parc_fts = parclot_lyr.getFeatures(QgsFeatureRequest(p_exp))
                    ws_name = f"Lot {lot}"
                    self.send_msg(exportxl_msg_txt[1].format(ws_name))
                    self.ws = wb.create_sheet(ws_name)
                    # Fills the header
                    hd_fld_exp = self.att_export[1:-1].split("','")
                    hd_text = self.hd_export.split(",")
                    nb_col = len(hd_text)
                    for col, hd_txt in enumerate(hd_text):
                        self.fill_cell(hd_txt, 1, col+1, fill_color="bbbbbb", b=True)
                    # Fills the rows (parcelle)
                    row = 1
                    for parc in parc_fts:
                        row += 1
                        for col, expr_val in enumerate(hd_fld_exp):
                            nw_val = get_val_by_expr(expr_val, parclot_lyr, parc)
                            self.fill_cell(nw_val, row , col + 1)
                    # Adds total row
                    row += 2
                    self.fill_cell(tot_label.format(lot), row , 1, fill_color="bbbbbb", b=True)
                    tot_cont = getmulti_fields_values_from_one_value(nwlot_lyr, lot, self.lot_attname, '', '')[0][tot_fldname]
                    self.fill_cell(transfo_m_to_ha(tot_cont), row , 2, fill_color="eeeeee")
                    # Calculates columns width
                    for column_cells in self.ws.columns:
                        length = max(len(as_text(cell.value)) for cell in column_cells) + col_delta_w
                        self.ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
                        
                wb.save(xl_filepath)
                open_file(xl_filepath)
                
                # End message
                self.send_msg(exportxl_msg_txt[2])
                QMessageBox.information(
                                        self.feedback, 
                                        exportxl_msg_txt[0], 
                                        exportxl_msg_txt[2]
                                        )

            except Exception as e:
                self.send_msg(alert_fic_msg_txt[1].format(str(e)))
                QMessageBox.warning(
                                        self.feedback, 
                                        alert_fic_msg_txt[0], 
                                        alert_fic_msg_txt[1].format(str(e))
                                        )
            self.feedback.close()


    def fill_cell(self, val, start_row, start_col, end_row=None, end_col=None, fill_color = "ffffff", b = False):
        '''
             Fills a cell or merged cells if end_row and end_col != None)
            # b: boolean for bold text
        '''
        cell = self.ws.cell(start_row, start_col, val)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(fill_type='solid',
                                fgColor=fill_color)
        brd = Side(style='thin', color='00000000')
        cell.border = Border(left=brd, right=brd,
                             top=brd, bottom=brd)
        cell.font = Font(bold=b)
        if end_col or end_row:
            if not end_row:
                end_row = start_row
            if not end_col:
                end_col = start_col
            self.ws.merge_cells(start_row=start_row,
                                start_column=start_col,
                                end_row=end_row,
                                end_column=end_col)


    def send_msg(self, msg):
        '''
            Send message to the dlg
        '''
        self.feedback.pg_bar.setValue(self.pgb_val)
        self.feedback.update_log(msg)
        self.pgb_val += 1


class MsgWnd(QWidget, gui_dlg_hunting_msg):
    '''
        Manages the message window
    '''
    send_ok = pyqtSignal()
    
    # Initialization
    def __init__(self, parent=None):
        super(MsgWnd, self).__init__(parent)
        self.setupUi(self)
        # Fills title and button text
        self.setWindowTitle(dlg_msg_titles[1])
        self.valid_bt.setText(dlg_msg_buts[1])
        # Initialization of the closing method (False = quit by red cross)
        self.quit_valid = False
        # Deletes Widget on close event
        self.setAttribute(Qt.WA_DeleteOnClose)
        # Connections
        self.valid_bt.clicked.connect(self.butt_ok)


    def butt_ok(self):
        '''
            Closes the window when clicking on the launch button
        '''
        self.quit_valid = True
        self.close()


    def closeEvent(self, event):
        '''
            Sends signal when the window is quit
        '''
        if self.quit_valid:
            self.send_ok.emit()
        else:
            # Hides the window
            self.hide()


    def update_log(self, msg):
        """
            Updates the message zone
        """
        t = self.msg_ted
        t.ensureCursorVisible()
        prefix = '<span style="font-weight:normal;">'
        suffix = '</span>'
        t.append('%s %s %s' % (prefix, msg, suffix))
        c = t.textCursor()
        c.movePosition(QTextCursor.End, QTextCursor.MoveAnchor)
        t.setTextCursor(c)
        qApp.processEvents()
